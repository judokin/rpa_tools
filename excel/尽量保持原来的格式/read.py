from openpyxl import load_workbook
import pandas as pd
import json

excel_file = "D:\\data\\凯燕归创建货件自动化测试.xlsx"
df = pd.read_excel(excel_file)
df.values
#import pdb;pdb.set_trace()

wb = load_workbook('ManifestFileUpload_Template_IncludeCasePack_IncludeExpirationDate_IncludeMLC_MPL.xlsx')
ws = wb.worksheets[2] 

# 修改某单元格的数据
ws['B3'] = 'Seller'
ws['B4'] = 'Seller'
for i, v in enumerate(df['SKU']):
    ws['A{}'.format(i + 9)] = v
for i, v in enumerate(df['数量']):
    ws['B{}'.format(i + 9)] = v 
    ws['C{}'.format(i + 9)] = 'Seller' 
    ws['D{}'.format(i + 9)] = 'Seller'
    #import pdb;pdb.set_trace()
    ws['I{}'.format(i + 9)] = str(df[df.keys()[11]].loc[i])
    ws['J{}'.format(i + 9)] = str(df[df.keys()[12]].loc[i])
    ws['K{}'.format(i + 9)] = str(df[df.keys()[13]].loc[i])
    ws['L{}'.format(i + 9)] = str(df[df.keys()[14]].loc[i])

# 每箱
for i, v in enumerate(df[df.keys()[10]]):
    ws['G{}'.format(i + 9)] = v
# 箱数
for i, v in enumerate(zip(df['数量'], df[df.keys()[10]])):
    ws['H{}'.format(i + 9)] = v[0]/v[1]
wb.save('modified.xlsx')