import json
#import config
import requests
import os
import time
import importlib.util
import pandas as pd
from openpyxl import Workbook
from openpyxl.drawing.image import Image as ExcelImage
from openpyxl.utils import get_column_letter

# 指定文件路径
config_file_path = r"D:\rpa_tools\feishu\config.py"

# 加载模块
spec = importlib.util.spec_from_file_location("config", config_file_path)
config = importlib.util.module_from_spec(spec)
spec.loader.exec_module(config)

url = "https://open.feishu.cn/open-apis/auth/v3/tenant_access_token/internal/"
payload = json.dumps({
"app_id": config.app_id,
"app_secret": config.app_secret
})
headers = {
'Content-Type': 'application/json'
}
response = requests.request("POST", url, headers=headers, data=payload)
print(response.status_code)
print(response.text)
tenant_access_token = response.json()["tenant_access_token"]

def download_file(file_name, url):
    headers = {
        'Authorization': 'Bearer ' + tenant_access_token
    }
    #url = "https://open.feishu.cn/open-apis/drive/v1/medias/batch_get_tmp_download_url?file_tokens=FM5ebnCXJo36FLx9ii4cMrsKn9c&extra=%7B%22bitablePerm%22%3A%7B%22tableId%22%3A%22tblux7wXHLPNgroJ%22%2C%22rev%22%3A3%7D%7D"
    #response = requests.request("GET", url, headers=headers, data=payload)
    response = requests.get(url, headers=headers, stream=True)

    if response.status_code == 200:
        # 将文件保存到本地
        # file_name = "test"
        with open(file_name, "wb") as file:
            for chunk in response.iter_content(chunk_size=8192):
                if chunk:  # 过滤掉保持活动的空块
                    file.write(chunk)
        print(f"文件已成功保存为: {file_name}")
    else:
        print(f"下载失败，状态码: {response.status_code}，响应: {response.text}")

def reset_fields(data, record_id):
    url = f"https://open.feishu.cn/open-apis/bitable/v1/apps/KP5ubEedLatObjs2EBccwPiEnSf/tables/tblWZWLW24tKmHyM/records/{record_id}"
    response = requests.put(url, headers=headers, json=data)

    print(response.status_code)
    print(response.json())

def date_range():
    import datetime
    # 获取当前时间
    now = datetime.datetime.now()
    # 获取当前时间后面10个整点时间
    next_10_hours = [
        (now + datetime.timedelta(hours=i+4)).replace(minute=0, second=0, microsecond=0)
        for i in range(1, 11)
    ]
    # 输出结果
    for dt in next_10_hours:
        print(dt)
    return next_10_hours
def get_table_data():
    #date_ranges = date_range()
    # 查询多维表格数据
    url = "https://open.feishu.cn/open-apis/bitable/v1/apps/KP5ubEedLatObjs2EBccwPiEnSf/tables/tbl91cqJf5rE7ANq/records/search?page_size=999"

    # 定义请求头
    headers = {
        "Authorization": "Bearer " + tenant_access_token,
        "Content-Type": "application/json"
    }

    # 将'YYYY-MM-DD HH:MM:SS'替换为一个实际的日期时间字符串
    # timestamp = int(time.mktime(time.strptime('2025-01-03 00:00:00', '%Y-%m-%d %H:%M:%S'))) * 1000
    # 定义请求体
    # data = {'fields': {'上传日期': timestamp}}

    data = {}
    # 发送 POST 请求
    response = requests.post(url, headers=headers, json=data)

    # 输出返回的状态码和响应数据
    print("Status Code:", response.status_code)
    data_json = {}
    try:
        data_json = response.json()
        #print("Response JSON:", response.json())
    except ValueError:
        pass
        #print("Response Text:", response.text)
    newest_day = ''
    break_times = 0
    update_datas_list = []
    datas = []
    for items in data_json['data']['items']:
        print(items)
        folder_path = r"D:\data\feishu_excel\img\\"
        if not os.path.exists(folder_path):
            os.makedirs(folder_path)
        if '后台截图' not in items['fields']:
            continue
        file_name = folder_path + items['fields']['后台截图'][0]['name']
        if not os.path.exists(file_name):
            download_file(file_name, items['fields']['后台截图'][0]['url'])
        item_dict = {}
        item_dict['FBA货件编号'] = items['fields']['FBA货件编号'][0]['text']
        item_dict['MSKU'] = items['fields']['MSKU'][0]['text']
        item_dict['AMAZON LINK'] = items['fields']['AMAZON LINK'][0]['text']
        item_dict['后台截图'] = items['fields']['后台截图'][0]['name'] # Corrected to use 'name' instead of 'text'
        item_dict['店铺'] = items['fields']['店铺'][0]['text']
        item_dict['销价登记'] = items['fields']['销价登记'][0]['text']
        item_dict['登记时间'] = items['fields']['登记时间'][0]['text']
        datas.append(item_dict)

    # Create a Pandas DataFrame from the datas list
    df = pd.DataFrame(datas)

    # Define the output directory and file path
    output_dir = r"D:\data\feishu_excel"
    if not os.path.exists(output_dir):
        os.makedirs(output_dir)
    excel_file_path = os.path.join(output_dir, f"feishu_data_with_images_{time.strftime('%Y%m%d_%H%M%S')}.xlsx")

    # Create a new Excel workbook and add a worksheet
    wb = Workbook()
    ws = wb.active
    ws.title = "Feishu Data"

    # Write headers
    headers = list(df.columns)
    ws.append(headers)

    # Write data and embed images
    for r_idx, row_data in df.iterrows():
        row_values = []
        for c_idx, header in enumerate(headers):
            if header == '后台截图':
                # Get the image file path
                image_name = row_data[header]
                image_path = os.path.join(folder_path, image_name)
                if os.path.exists(image_path):
                    # Insert image into the cell
                    img = ExcelImage(image_path)
                    # Calculate appropriate size and position for the image
                    # You might need to adjust these values based on your image sizes
                    ws.row_dimensions[r_idx + 2].height = 100 # Adjust row height
                    ws.column_dimensions[get_column_letter(c_idx + 1)].width = 20 # Adjust column width
                    img.width = 120
                    img.height = 90
                    ws.add_image(img, f'{get_column_letter(c_idx + 1)}{r_idx + 2}')
                    row_values.append(f"See image: {image_name}") # Add a placeholder text
                else:
                    row_values.append(f"Image not found: {image_name}")
            else:
                row_values.append(row_data[header])
        ws.append(row_values)

    # Save the workbook
    wb.save(excel_file_path)
    print(f"Data successfully saved to {excel_file_path}")


if __name__ == "__main__":
    get_table_data()
