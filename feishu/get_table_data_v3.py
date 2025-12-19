import json
import requests
import os
import time
import importlib.util
import configparser
import datetime
import hashlib
import urllib.parse
from Crypto.Cipher import AES
from Crypto.Util.Padding import pad
import base64
from datetime import timedelta, datetime
import pandas as pd
import openpyxl
from openpyxl.utils.dataframe import dataframe_to_rows

'''
[DEFAULT]
appSecret:ysy0Gtn1pDRdQTGshcl1fA==
appId:ak_rhvvA99g8tNpb
'''
APP_SECRET = "ysy0Gtn1pDRdQTGshcl1fA=="
APP_ID = "ak_rhvvA99g8tNpb"
# 指定文件路径
config_file_path = r"config.py"

# 加载模块
spec = importlib.util.spec_from_file_location("config", config_file_path)
config = importlib.util.module_from_spec(spec)
spec.loader.exec_module(config)

# Load Lingxing API config from config.ini
lingxing_config_parser = configparser.ConfigParser()
lingxing_config_parser.read('config.ini')

# Feishu tenant access token
url = "https://open.feishu.cn/open-apis/auth/v3/tenant_access_token/internal/"
payload = json.dumps({
"app_id": config.app_id,
"app_secret": config.app_secret
})
headers = {
'Content-Type': 'application/json'
}
response = requests.request("POST", url, headers=headers, data=payload)
print(response.status_code)
print(response.text)
tenant_access_token = response.json()["tenant_access_token"]


# appSecret:ysy0Gtn1pDRdQTGshcl1fA==
# appId:ak_rhvvA99g8tNpb
# Functions from p.py for Lingxing API authentication
def get_access_token_lingxing():
    url = 'https://openapi.lingxing.com/api/auth-server/oauth/access-token'
    data = {
        'appSecret': 'ysy0Gtn1pDRdQTGshcl1fA==',
        'appId': 'ak_rhvvA99g8tNpb'
    }
    response = requests.post(url, data=data)
    if response.status_code == 200:
        return response.json()
    else:
        print(f"Lingxing Access Token Request failed with status code {response.status_code}")
        print("Response content:", response.text)
        return None

def generate_sign(params, access_token, timestamp, app_id):
    """
    生成签名并进行AES加密和URL编码。

    :param params: dict, 业务请求参数
    :param access_token: str, 访问令牌G
    :param timestamp: str, 时间戳
    :param app_id: str, 应用ID，用作AES加密密钥
    :return: dict, 包含生成的签名和编码后的签名
    """
    # 添加固定参数
    params.update({
        'access_token': access_token,
        'app_key': app_id,
        'timestamp': timestamp
    })
    # 按照key的ASCII排序
    sorted_params = sorted(params.items(), key=lambda item: item)

    # 拼接参数，忽略值为空的参数
    param_str = '&'.join([f"{k}={v}" for k, v in sorted_params if v not in [None, '']])
    param_str = param_str.replace(" ", "")
    param_str = param_str.replace("'", '"')
    print("param_str =", param_str)
    # 生成MD5签名并转换为大写
    md5_hash = hashlib.md5(param_str.encode('utf-8')).hexdigest().upper()

    # AES加密配置
    key = app_id.encode('utf-8')  # AES密钥需要为bytes
    cipher = AES.new(key, AES.MODE_ECB)

    # 加密MD5值，填充PKCS5
    encrypted = cipher.encrypt(pad(md5_hash.encode('utf-8'), AES.block_size))

    # 将加密后的字节转为Base64编码字符串
    encrypted_sign = base64.b64encode(encrypted).decode('utf-8')

    # URL编码签名
    url_encoded_sign = urllib.parse.quote(encrypted_sign)

    return {
        'sign': md5_hash,
        'encrypted_sign': url_encoded_sign
    }

def generate_sign_lingxing(params, access_token, timestamp, app_id):
    # Create a copy to avoid modifying the original params dictionary
    params_for_signing = params.copy()

    # Add fixed parameters
    # params_for_signing.update({
    #     'access_token': access_token,
    #     'app_key': app_id,
    #     'timestamp': timestamp
    # })

    # Serialize complex objects (like lists/dicts) to JSON strings for signing
    serialized_params = {}
    for k, v in params_for_signing.items():
        if isinstance(v, (list, dict)):
            serialized_params[k] = json.dumps(v, ensure_ascii=False)
        else:
            serialized_params[k] = v

    # Sort parameters by key's ASCII value
    sorted_params = sorted(serialized_params.items(), key=lambda item: item[0])

    # Concatenate parameters, ignoring empty values
    param_str = '&'.join([f"{k}={v}" for k, v in sorted_params if v not in [None, '']])
    param_str = param_str.replace("'", '"')
    #import json
    #import pdb;pdb.set_trace()
    print("params ==== ", json.dumps(params))
    print("param_str =", param_str)
    # Generate MD5 signature and convert to uppercase
    md5_hash = hashlib.md5(param_str.encode('utf-8')).hexdigest().upper()

    # AES encryption configuration
    key = app_id.encode('utf-8')
    cipher = AES.new(key, AES.MODE_ECB)

    # Encrypt MD5 value, pad with PKCS5
    encrypted = cipher.encrypt(pad(md5_hash.encode('utf-8'), AES.block_size))

    # Convert encrypted bytes to Base64 string
    encrypted_sign = base64.b64encode(encrypted).decode('utf-8')

    # URL encode signature
    url_encoded_sign = urllib.parse.quote(encrypted_sign)

    return {
        'sign': md5_hash,
        'encrypted_sign': url_encoded_sign
    }

def reset_fields(data, record_id):
    url = f"https://open.feishu.cn/open-apis/bitable/v1/apps/KP5ubEedLatObjs2EBccwPiEnSf/tables/tblWZWLW24tKmHyM/records/{record_id}"
    response = requests.put(url, headers=headers, json=data)

    print(response.status_code)
    print(response.json())
def excute_sql(sql):
    import os
    import json
    import datetime
    import mysql.connector
    from mysql.connector import Error
    import sys
    import concurrent.futures

    # 配置数据库信息
    DB_CONFIG = {
        "host": "116.63.47.216",
        "user": "zhuzhenjin",
        "password": "plORmnE*KYnQRvF8",
        "database": "luyao"
    }
    """执行一个批次的SQL语句，每batch_size条提交一次"""
    try:
        connection = mysql.connector.connect(**DB_CONFIG)
        cursor = connection.cursor()
        cursor.execute(sql)
        connection.commit()  # 批量提交
        print("SQL语句已成功执行！")
    except Error as e:
        message = f"数据库错误: {e}"
        print(message)
        connection.rollback()
    finally:
        if connection.is_connected():
            cursor.close()
            connection.close()
def submit_pricing_changes(pricing_params_list):
    api_url = "https://openapi.lingxing.com/erp/sc/listing/ProductPricing/pricingSubmit"
    
    access_token_info = get_access_token_lingxing()
    if not access_token_info or 'data' not in access_token_info:
        print("Failed to get Lingxing access token.")
        return

    access_token = access_token_info['data']['access_token']
    try:
        app_id = lingxing_config_parser['DEFAULT']['appId']
    except:
        app_id = APP_ID
    timestamp = str(int(time.time()))

    # The actual request body, which will also be used for signing
    request_body_for_api = {
        "pricing_params": pricing_params_list
        # Add fixed parameters to the request body as per p.py's signing method
        # "access_token": access_token,
        # "app_key": app_id,
        # "timestamp": timestamp
    }

    # Pass the *entire* request_body_for_api to generate_sign_lingxing.
    # The generate_sign_lingxing function will then sort and serialize this dictionary.
    # It will also add access_token, app_key, timestamp again, but since they are already there,
    # the update will just overwrite them with the same values, which is harmless.
    sign_info = generate_sign(request_body_for_api, access_token, timestamp, app_id)
    
    # The URL query string still needs the fixed parameters and the generated sign
    ext_url = f"access_token={access_token}&app_key={app_id}&timestamp={timestamp}&sign={sign_info['encrypted_sign']}"
    full_url = f"{api_url}?{ext_url}"

    headers = {
        'Content-Type': 'application/json'
    }

    print(f"Submitting pricing changes to: {full_url}")
    # Print the request_body_for_api, which now includes the fixed parameters
    print(f"Payload: {json.dumps(request_body_for_api, indent=2, ensure_ascii=False)}")

    try:
        # Use the *updated* request_body_for_api as the JSON body
        response = requests.post(full_url, headers=headers, json=request_body_for_api, timeout=60)
        if response.status_code == 200:
            return response.json()
        else:
            print(f"Failed to submit pricing changes: Status Code {response.status_code}, Response: {response.text}")
            return False
    except requests.exceptions.RequestException as e:
        print(f"An error occurred while submitting pricing changes: {e}")
        return False


def get_table_data():
    # https://wit0jhu6kvu.feishu.cn/base/IV21b4ZnDaNHpOs3cDYcgDL5nqe?table=tblqBhJFrkDCqTHU&view=vew4KhjGpe
    # https://wit0jhu6kvu.feishu.cn/base/IV21b4ZnDaNHpOs3cDYcgDL5nqe?table=tblqBhJFrkDCqTHU&view=vew4KhjGpe
    # https://wit0jhu6kvu.feishu.cn/base/IV21b4ZnDaNHpOs3cDYcgDL5nqe?table=tbl6MNmYcyk5uoqt&view=vewITPRIVy
    # 查询多维表格数据
    url = "https://open.feishu.cn/open-apis/bitable/v1/apps/IV21b4ZnDaNHpOs3cDYcgDL5nqe/tables/tbl6MNmYcyk5uoqt/records/search?page_size=999"

    # 定义请求头
    headers = {
        "Authorization": "Bearer " + tenant_access_token,
        "Content-Type": "application/json"
    }
    data = {}
    # 发送 POST 请求
    response = requests.post(url, headers=headers, json=data)

    # 输出返回的状态码和响应数据
    print("Status Code:", response.status_code)
    data_json = {}
    try:
        data_json = response.json()
        #print("Response JSON:", response.json())
    except ValueError:
        pass
        #print("Response Text:", response.text)
    print(data_json)

    # Data to be saved to Excel
    excel_data = []
    sql_data = []
    # header 为 *MSKU	*店铺名称	*国家	价格	价格积分(%)	优惠价	优惠价开始日期	优惠价结束日期	优惠价积分(%)
    excel_headers = ['MSKU', '店铺名称', '国家', '价格', '价格积分(%)', '优惠价', '优惠价开始日期', '优惠价结束日期', '优惠价积分(%)']

    for index, items in enumerate(data_json['data']['items']):
        print("index =", index)
        if '运营审核' in items['fields'] and items['fields']['运营审核'] == '不通过':
            continue
        print(items)
        
        fields = items['fields']
        
        msku = fields.get('MSKU')[0]['text']
        shop_name = fields.get('店铺', '')
        site = fields.get('站点')[0]['text'] if fields.get('站点') else ''
        standard_price = fields.get('价格', '')
        price_points_percentage = fields.get('价格积分(%)', '')
        sale_price = fields.get('待调优惠价', '')
        # sale_start_date 为当天的前一天
        sale_start_date = (datetime.now() - timedelta(days=1)).strftime("%Y/%m/%d")
        # sale_end_date 为当天的10年后
        sale_end_date = (datetime.now() + timedelta(days=3650)).strftime("%Y/%m/%d")
        sale_points_percentage = fields.get('优惠价积分(%)', '')

        excel_data.append({
            'MSKU': msku,
            '店铺名称': shop_name,
            '国家': site,
            '价格': standard_price if float(standard_price) >= float(sale_price) else float(sale_price),
            '价格积分(%)': price_points_percentage,
            '优惠价': sale_price,
            '优惠价开始日期': sale_start_date,
            '优惠价结束日期': sale_end_date,
            '优惠价积分(%)': sale_points_percentage,
        })
        sql_data.append({
            'msku': msku,
            'shop_name': shop_name,
            'sid': fields.get('店铺ID', '4188'),
            'site': site,
            'sale_price': standard_price if float(standard_price) >= float(sale_price) else float(sale_price),
            'sale_price': sale_price,
            'sale_start_date': sale_start_date,
            'sale_end_date': sale_end_date,
            'asin': fields.get('ASIN')[0]['text'],
            'record_id': items.get('record_id')
        })

    if excel_data:
        df_output = pd.DataFrame(excel_data, columns=excel_headers)
        
        template_path = r"D:\data\lingxing\导入调价模板.xlsx"
        today_date = datetime.now().strftime("%Y%m%d")
        output_filename = f"feishu_table_data_{today_date}.xlsx"
        output_path = os.path.join(r"D:\data\lingxing", output_filename)

        try:
            # Load the template workbook
            book = openpyxl.load_workbook(template_path)
            sheet = book.active # Or specify sheet name: book['Sheet1']

            # Clear existing data rows below the header (assuming header is in row 1)
            # This clears all rows from row 2 onwards
            for row in sheet.iter_rows(min_row=2):
                for cell in row:
                    cell.value = None
            
            # Write the new DataFrame data to the sheet, starting from row 2
            # header=False to not write DataFrame headers again
            # startrow=1 because openpyxl is 1-indexed for rows, and we want to start at the second row
            for r_idx, row in enumerate(dataframe_to_rows(df_output, index=False, header=False), 2):
                for c_idx, value in enumerate(row, 1):
                    sheet.cell(row=r_idx, column=c_idx, value=value)

            # Save the modified workbook to a new file
            book.save(output_path)
            print(f"Data successfully saved to {output_path} while preserving template format.")
        except FileNotFoundError:
            print(f"Error: Template file not found at {template_path}. Creating a new Excel file.")
            df_output.to_excel(output_path, index=False)
            print(f"Data saved to new file {output_path}")
        except Exception as e:
            print(f"An error occurred while saving to Excel: {e}")
            # Fallback to saving as a new file if template manipulation fails
            df_output.to_excel(output_path, index=False)
            print(f"Data saved to new file {output_path} due to error.")
    else:
        print("No data to save to Excel.")

    # New logic to insert sql_data into the database
    if sql_data:
        table_name = "ods_lingxing_set_price" # Assuming a table name
        for record in sql_data:
            columns = ', '.join(record.keys())
            values_list = []
            for v in record.values():
                if isinstance(v, str):
                    # Escape single quotes and enclose in single quotes for SQL
                    v = v.replace("'", "''")
                    values_list.append(f"'{v}'")
                else:
                    values_list.append(str(v)) # Convert non-string values directly
            values = ', '.join(values_list)
            sql_insert_statement = f"INSERT IGNORE INTO {table_name} ({columns}) VALUES ({values})"
            print(f"Executing SQL: {sql_insert_statement}")
            excute_sql(sql_insert_statement)
    else:
        print("No SQL data to insert into the database.")


if __name__ == "__main__":
    get_table_data()
